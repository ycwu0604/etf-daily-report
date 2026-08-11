"""
ETF 資料處理器
- 解析 ezmoney 下載的 xlsx → 寫入 SQLite
- 產出每日比較報表（A）+ 歷史股數展開（C）

使用方式：
  python etf_processor.py                          # 處理 downloads/ 下所有 xlsx
  python etf_processor.py --file path/to/file.xlsx # 處理指定檔案
  python etf_processor.py --report                 # 只產報表（不重新解析）
"""

import sqlite3
import re
import sys
import openpyxl
from pathlib import Path
from datetime import datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ====== 路徑設定 ======
PROJECT_DIR = Path(__file__).parent
DB_PATH = PROJECT_DIR / "etf_data.db"
DOWNLOADS_DIR = PROJECT_DIR / "downloads"
REPORT_DIR = PROJECT_DIR / "reports"

# ====== 民國日期 → 西元 ======
def roc_to_ad(roc_str: str) -> str:
    """'115/07/09' → '2026-07-09'"""
    m = re.match(r"(\d+)/(\d+)/(\d+)", roc_str)
    if not m:
        return roc_str
    year = int(m.group(1)) + 1911
    return f"{year}-{m.group(2)}-{m.group(3)}"


# ====== 解析 NTD 金額字串 ======
def parse_amount(val) -> float | None:
    """'NTD 274,753,577,961' → 274753577961.0"""
    if val is None:
        return None
    s = str(val).replace("NTD", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


# ====== 解析百分比 ======
def parse_pct(val) -> float | None:
    """'97.08%' → 97.08"""
    if val is None:
        return None
    s = str(val).replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


# ============================================================
#  SQLite Schema
# ============================================================
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS daily_summary (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    net_assets  REAL,
    nav         REAL,
    outstanding_units REAL,
    stock_amount REAL,
    cash_amount  REAL,
    futures_margin REAL,
    repo_bonds   REAL,
    receivables  REAL,
    PRIMARY KEY (date, fund_code)
);

CREATE TABLE IF NOT EXISTS daily_holdings (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    stock_code  TEXT NOT NULL,
    stock_name  TEXT NOT NULL,
    shares      INTEGER,
    weight_pct  REAL,
    PRIMARY KEY (date, fund_code, stock_code)
);

CREATE TABLE IF NOT EXISTS daily_futures (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    future_code TEXT NOT NULL,
    future_name TEXT NOT NULL,
    weight_pct  REAL,
    lots        INTEGER,
    contract_month TEXT,
    PRIMARY KEY (date, fund_code, future_code, contract_month)
);
"""


def init_db(db: sqlite3.Connection):
    db.executescript(SCHEMA_SQL)
    db.commit()


# ============================================================
#  解析 xlsx → 寫入 DB
# ============================================================
def parse_xlsx(filepath: Path, db: sqlite3.Connection, fund_code_override: str | None = None) -> bool:
    """解析一個 xlsx，回傳 (成功與否)"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # --- 從檔名推導 fund_code ---
    fname = filepath.stem  # e.g. '49YTW_ETF_Investment_Portfolio_20260709'
    fc_match = re.match(r"^([A-Z0-9]{4,6})_", fname)
    if fc_match:
        fund_code = fc_match.group(1)
    elif fund_code_override:
        fund_code = fund_code_override
    else:
        fund_code = "UNKNOWN"

    # --- 日期 ---
    raw_date = ws["A1"].value  # '資料日期：115/07/09'
    date_str = roc_to_ad(raw_date.replace("資料日期：", "").replace("資料日期:", "").strip()) if raw_date else ""

    if not date_str or not re.match(r"\d{4}-\d{2}-\d{2}", date_str):
        print(f"  ❌ 無法解析日期：{raw_date}")
        wb.close()
        return False

    # --- 檢查是否已有持股資料 ---
    existing = db.execute(
        "SELECT 1 FROM daily_holdings WHERE date=? AND fund_code=? LIMIT 1",
        (date_str, fund_code),
    ).fetchone()
    if existing:
        print(f"  ⏭️  [{fund_code}] {date_str} 資料已存在，跳過")
        wb.close()
        return True

    # --- 基金摘要 ---
    net_assets = parse_amount(ws["B4"].value)
    outstanding_units = parse_amount(ws["B5"].value)
    nav = parse_amount(ws["B6"].value)
    stock_amount = parse_amount(ws["B10"].value)
    cash_amount = parse_amount(ws["B13"].value)
    futures_margin = parse_amount(ws["B14"].value)
    repo_bonds = parse_amount(ws["B15"].value)
    receivables = parse_amount(ws["B16"].value)

    db.execute(
        """INSERT OR REPLACE INTO daily_summary
           (date, fund_code, net_assets, nav, outstanding_units,
            stock_amount, cash_amount, futures_margin, repo_bonds, receivables)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (date_str, fund_code, net_assets, nav, outstanding_units,
         stock_amount, cash_amount, futures_margin, repo_bonds, receivables),
    )

    # --- 期貨明細 ---
    # 從 xlsx 動態搜尋「期貨代號」表頭行，而非寫死 row number
    # 格式: 期貨代號(A) | 期貨名稱(B) | 持股權重(C) | 口數(D) | 契約年月(E)
    futures_header_row = None
    futures_rows = []
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() in ('期貨代號', '期�貨代號'):
            futures_header_row = row
            break

    if futures_header_row:
        for row in range(futures_header_row + 1, ws.max_row + 1):
            code = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            if not code or not name:
                break  # 期貨區塊結束
            weight_raw = ws.cell(row=row, column=3).value
            lots_raw = ws.cell(row=row, column=4).value
            contract = ws.cell(row=row, column=5).value

            weight = parse_pct(weight_raw)
            try:
                lots = int(str(lots_raw).replace(",", "")) if lots_raw else 0
            except ValueError:
                lots = 0

            futures_rows.append((
                date_str, fund_code, str(code).strip(), str(name).strip(),
                weight, lots, str(contract).strip() if contract else None,
            ))

    if futures_rows:
        # 檢查是否已有期貨資料
        existing_futures = db.execute(
            "SELECT 1 FROM daily_futures WHERE date=? AND fund_code=? LIMIT 1",
            (date_str, fund_code),
        ).fetchone()
        if not existing_futures:
            db.executemany(
                """INSERT OR REPLACE INTO daily_futures
                   (date, fund_code, future_code, future_name, weight_pct, lots, contract_month)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                futures_rows,
            )
            print(f"  📈 [{fund_code}] {date_str} 寫入 {len(futures_rows)} 筆期貨")

    # --- 持股明細 ---
    # 動態搜尋「股票代號」表頭行，避免因期貨有無導致偏移
    holdings_header_row = None
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val and str(val).strip() in ('股票代號', '股�票代號'):
            holdings_header_row = row
            break

    start_row = (holdings_header_row + 1) if holdings_header_row else 21
    holdings = []
    for row in range(start_row, ws.max_row + 1):
        stock_code = ws.cell(row=row, column=1).value
        stock_name = ws.cell(row=row, column=2).value
        shares_raw = ws.cell(row=row, column=3).value
        weight_raw = ws.cell(row=row, column=4).value

        if not stock_code or not stock_name:
            continue

        # 股數：可能是數字或帶逗號的字串
        try:
            shares = int(str(shares_raw).replace(",", "")) if shares_raw else 0
        except ValueError:
            shares = 0

        weight = parse_pct(weight_raw)

        holdings.append((date_str, fund_code, str(stock_code), str(stock_name), shares, weight))

    db.executemany(
        """INSERT OR REPLACE INTO daily_holdings
           (date, fund_code, stock_code, stock_name, shares, weight_pct)
           VALUES (?, ?, ?, ?, ?, ?)""",
        holdings,
    )

    db.commit()
    wb.close()
    print(f"  ✅ [{fund_code}] {date_str} 寫入 {len(holdings)} 筆持股")
    return True


def process_all_xlsx(db: sqlite3.Connection, downloads_dir: Path, fund_code_map: dict | None = None):
    """掃描 downloads/ 處理所有 xlsx
    fund_code_map: {filename_pattern: fund_code} 用於舊格式檔名
    """
    xlsx_files = sorted(downloads_dir.glob("*.xlsx"))
    xlsx_files = [f for f in xlsx_files if not f.name.startswith("~$")]

    if not xlsx_files:
        print("downloads/ 下沒有 xlsx 檔案")
        return

    fc_map = fund_code_map or {}
    print(f"找到 {len(xlsx_files)} 個 xlsx 檔案")
    for f in xlsx_files:
        print(f"\n--- 解析 {f.name} ---")
        # 先嘗試從檔名前綴提取 fund_code（新格式：49YTW_xxx.xlsx）
        fname = f.stem
        fc_match = re.match(r"^([A-Z0-9]{4,6})_", fname)
        if fc_match:
            # 檔名已有 fund_code 前綴，直接用
            parse_xlsx(f, db)
        else:
            # 舊格式檔名，用映射表
            override = None
            for pattern, code in fc_map.items():
                if pattern in f.name:
                    override = code
                    break
            parse_xlsx(f, db, fund_code_override=override)


# ============================================================
#  報表產出：每個 ETF 一個 sheet
#  左側：日間比較（今日 vs 昨日，股數增減）
#  右側：歷史股數按日展開
# ============================================================
def generate_report(db: sqlite3.Connection, output_path: Path):
    """產出每日比較 + 歷史股數報表"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # 取得所有 fund_code
    fund_codes = [r[0] for r in db.execute(
        "SELECT DISTINCT fund_code FROM daily_holdings ORDER BY fund_code"
    ).fetchall()]

    # 取得所有日期（排序）
    all_dates = [r[0] for r in db.execute(
        "SELECT DISTINCT date FROM daily_holdings ORDER BY date"
    ).fetchall()]

    if not fund_codes or not all_dates:
        print("資料庫中沒有資料，無法產出報表")
        return

    today = all_dates[-1]
    yesterday = all_dates[-2] if len(all_dates) >= 2 else None

    wb = Workbook()
    # 刪掉預設 sheet
    wb.remove(wb.active)

    # --- 樣式定義 ---
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    new_fill = PatternFill("solid", fgColor="C6EFCE")  # 新增：淺綠
    del_fill = PatternFill("solid", fgColor="FFC7CE")  # 刪除：淺紅
    increase_font = Font(color="006100")  # 增加：深綠
    decrease_font = Font(color="9C0006")  # 減少：深紅
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for fund_code in fund_codes:
        ws = wb.create_sheet(title=fund_code)

        # ====== 左側：日間比較 ======
        col_offset = 1  # 從 A 欄開始

        # 標題列
        ws.cell(row=1, column=col_offset, value=f"{fund_code} 每日持股比較").font = Font(bold=True, size=14)
        ws.cell(row=2, column=col_offset, value=f"日期：{today} vs {yesterday or 'N/A'}").font = Font(italic=True, size=10)

        # 表頭
        headers_a = ["股票代號", "股票名稱", f"股數 ({yesterday or 'N/A'})", f"股數 ({today})", "變化", "變化%"]
        header_row = 4
        for i, h in enumerate(headers_a):
            cell = ws.cell(row=header_row, column=col_offset + i, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 取得今日持股 {stock_code: (stock_name, shares)}
        today_holdings = {}
        for r in db.execute(
            "SELECT stock_code, stock_name, shares FROM daily_holdings WHERE date=? AND fund_code=? ORDER BY shares DESC",
            (today, fund_code),
        ).fetchall():
            today_holdings[r[0]] = (r[1], r[2])

        # 取得昨日持股
        yesterday_holdings = {}
        if yesterday:
            for r in db.execute(
                "SELECT stock_code, stock_name, shares FROM daily_holdings WHERE date=? AND fund_code=? ORDER BY shares DESC",
                (yesterday, fund_code),
            ).fetchall():
                yesterday_holdings[r[0]] = (r[1], r[2])

        # 合併所有股票代號（以今日為主，加入昨日有但今日沒有的）
        all_stocks = set(today_holdings.keys()) | set(yesterday_holdings.keys())
        # 排序：以今日股數降序，新增/刪除放最後
        def sort_key(code):
            t_shares = today_holdings.get(code, (None, 0))[1]
            y_shares = yesterday_holdings.get(code, (None, 0))[1]
            max_shares = max(t_shares, y_shares)
            is_new = code not in yesterday_holdings
            is_del = code not in today_holdings
            return (is_new or is_del, -max_shares)

        sorted_stocks = sorted(all_stocks, key=sort_key)

        data_row = header_row + 1
        for code in sorted_stocks:
            t_name, t_shares = today_holdings.get(code, ("", None))
            y_name, y_shares = yesterday_holdings.get(code, ("", None))
            name = t_name or y_name or code

            is_new = y_shares is None and t_shares is not None
            is_del = t_shares is None and y_shares is not None

            # 股數
            y_val = y_shares if y_shares is not None else "(無)"
            t_val = t_shares if t_shares is not None else "(無)"

            # 變化
            if is_new:
                change = "NEW"
                change_pct = ""
            elif is_del:
                change = "DEL"
                change_pct = ""
            else:
                diff = t_shares - y_shares
                change = diff
                change_pct = f"{diff / y_shares * 100:.2f}%" if y_shares != 0 else "N/A"

            ws.cell(row=data_row, column=col_offset, value=code).border = thin_border
            ws.cell(row=data_row, column=col_offset + 1, value=name).border = thin_border
            ws.cell(row=data_row, column=col_offset + 2, value=y_val).border = thin_border
            ws.cell(row=data_row, column=col_offset + 2).number_format = '#,##0'
            ws.cell(row=data_row, column=col_offset + 3, value=t_val).border = thin_border
            ws.cell(row=data_row, column=col_offset + 3).number_format = '#,##0'

            change_cell = ws.cell(row=data_row, column=col_offset + 4, value=change)
            change_cell.border = thin_border
            if isinstance(change, int):
                change_cell.number_format = '#,##0'
            pct_cell = ws.cell(row=data_row, column=col_offset + 5, value=change_pct)
            pct_cell.border = thin_border

            # 顏色標示
            if is_new:
                for c in range(col_offset, col_offset + 6):
                    ws.cell(row=data_row, column=c).fill = new_fill
            elif is_del:
                for c in range(col_offset, col_offset + 6):
                    ws.cell(row=data_row, column=c).fill = del_fill
            elif isinstance(change, int) and change > 0:
                change_cell.font = increase_font
                pct_cell.font = increase_font
            elif isinstance(change, int) and change < 0:
                change_cell.font = decrease_font
                pct_cell.font = decrease_font

            data_row += 1

        # ====== 右側：歷史股數按日展開 ======
        hist_col_offset = col_offset + len(headers_a) + 2  # 留 2 欄空白
        hist_header_row = 4

        ws.cell(row=1, column=hist_col_offset, value="歷史股數變化").font = Font(bold=True, size=14)

        # 表頭：股票代號 | 股票名稱 | 日期1 | 日期2 | ...
        ws.cell(row=hist_header_row, column=hist_col_offset, value="股票代號").font = header_font_white
        ws.cell(row=hist_header_row, column=hist_col_offset).fill = header_fill
        ws.cell(row=hist_header_row, column=hist_col_offset).border = thin_border
        ws.cell(row=hist_header_row, column=hist_col_offset + 1, value="股票名稱").font = header_font_white
        ws.cell(row=hist_header_row, column=hist_col_offset + 1).fill = header_fill
        ws.cell(row=hist_header_row, column=hist_col_offset + 1).border = thin_border

        for i, d in enumerate(all_dates):
            cell = ws.cell(row=hist_header_row, column=hist_col_offset + 2 + i, value=d)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 取得今日所有持股代號（以今日為主，也包含過去有但今日沒有的）
        all_ever_stocks = set()
        for d in all_dates:
            for row in db.execute(
                "SELECT DISTINCT stock_code FROM daily_holdings WHERE date=? AND fund_code=?",
                (d, fund_code),
            ).fetchall():
                all_ever_stocks.add(row[0])

        # 以今日股數排序
        def hist_sort_key(code):
            t = today_holdings.get(code, (None, 0))
            return -t[1]

        sorted_hist = sorted(all_ever_stocks, key=hist_sort_key)

        hist_row = hist_header_row + 1
        for code in sorted_hist:
            # 取名稱（優先取今日的）
            name = today_holdings.get(code, (code, 0))[0]
            if not name or name == code:
                for d in all_dates:
                    r = db.execute(
                        "SELECT stock_name FROM daily_holdings WHERE date=? AND fund_code=? AND stock_code=?",
                        (d, fund_code, code),
                    ).fetchone()
                    if r:
                        name = r[0]
                        break

            ws.cell(row=hist_row, column=hist_col_offset, value=code).border = thin_border
            ws.cell(row=hist_row, column=hist_col_offset + 1, value=name).border = thin_border

            # 先收集該股所有日期的股數，用來做 n vs n-1 比對
            hist_values = []
            for i, d in enumerate(all_dates):
                r = db.execute(
                    "SELECT shares FROM daily_holdings WHERE date=? AND fund_code=? AND stock_code=?",
                    (d, fund_code, code),
                ).fetchone()
                hist_values.append(r[0] if r else None)

            for i, val in enumerate(hist_values):
                cell = ws.cell(row=hist_row, column=hist_col_offset + 2 + i, value=val)
                cell.number_format = '#,##0'
                cell.border = thin_border
                if val is None:
                    cell.value = "-"
                    cell.font = Font(color="AAAAAA")
                else:
                    # 找前一個有效值（跳過 None）
                    prev_val = None
                    for j in range(i - 1, -1, -1):
                        if hist_values[j] is not None:
                            prev_val = hist_values[j]
                            break
                    if prev_val is not None and val != prev_val:
                        cell.fill = new_fill if val > prev_val else del_fill

            hist_row += 1

        # ====== 欄寬調整 ======
        ws.column_dimensions[get_column_letter(col_offset)].width = 10   # 股票代號
        ws.column_dimensions[get_column_letter(col_offset + 1)].width = 14  # 股票名稱
        ws.column_dimensions[get_column_letter(col_offset + 2)].width = 16  # 昨日股數
        ws.column_dimensions[get_column_letter(col_offset + 3)].width = 16  # 今日股數
        ws.column_dimensions[get_column_letter(col_offset + 4)].width = 12  # 變化
        ws.column_dimensions[get_column_letter(col_offset + 5)].width = 10  # 變化%
        ws.column_dimensions[get_column_letter(hist_col_offset)].width = 10
        ws.column_dimensions[get_column_letter(hist_col_offset + 1)].width = 14
        for i in range(len(all_dates)):
            ws.column_dimensions[get_column_letter(hist_col_offset + 2 + i)].width = 14

    # ====== 跨 ETF 合計 sheet ======
    if len(fund_codes) >= 2:
        summary_fill = PatternFill("solid", fgColor="FFC000")  # 橘色表頭
        summary_font = Font(bold=True, size=11, color="FFFFFF")

        # --- Sheet 1: 合計（跨 ETF 股數相加，按日展開）---
        ws_sum = wb.create_sheet(title="跨ETF合計")
        ws_sum.cell(row=1, column=1, value="跨 ETF 持股合計（股數相加）").font = Font(bold=True, size=14)
        ws_sum.cell(row=2, column=1, value=f"包含：{', '.join(fund_codes)}").font = Font(italic=True, size=10)

        # 表頭：股票代號 | 股票名稱 | 日期1 | 日期2 | ...
        sum_hdr_row = 4
        ws_sum.cell(row=sum_hdr_row, column=1, value="股票代號").font = summary_font
        ws_sum.cell(row=sum_hdr_row, column=1).fill = summary_fill
        ws_sum.cell(row=sum_hdr_row, column=1).border = thin_border
        ws_sum.cell(row=sum_hdr_row, column=2, value="股票名稱").font = summary_font
        ws_sum.cell(row=sum_hdr_row, column=2).fill = summary_fill
        ws_sum.cell(row=sum_hdr_row, column=2).border = thin_border

        for i, d in enumerate(all_dates):
            cell = ws_sum.cell(row=sum_hdr_row, column=3 + i, value=d)
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # 收集所有股票代號（跨所有 ETF、所有日期）
        all_cross_stocks = set()
        for r in db.execute(
            "SELECT DISTINCT stock_code FROM daily_holdings"
        ).fetchall():
            all_cross_stocks.add(r[0])

        # 以「最新日期的合計股數」降序排列
        latest_totals = {}
        for code in all_cross_stocks:
            total = 0
            for r in db.execute(
                "SELECT shares FROM daily_holdings WHERE date=? AND stock_code=?",
                (today, code),
            ).fetchall():
                total += r[0] if r[0] else 0
            latest_totals[code] = total

        sorted_cross = sorted(all_cross_stocks, key=lambda c: -latest_totals.get(c, 0))

        sum_row = sum_hdr_row + 1
        for code in sorted_cross:
            # 取名稱（優先最新日期）
            name = ""
            for fc in fund_codes:
                r = db.execute(
                    "SELECT stock_name FROM daily_holdings WHERE date=? AND fund_code=? AND stock_code=?",
                    (today, fc, code),
                ).fetchone()
                if r:
                    name = r[0]
                    break
            if not name:
                for d in all_dates:
                    for fc in fund_codes:
                        r = db.execute(
                            "SELECT stock_name FROM daily_holdings WHERE date=? AND fund_code=? AND stock_code=?",
                            (d, fc, code),
                        ).fetchone()
                        if r:
                            name = r[0]
                            break
                    if name:
                        break
            if not name:
                name = code

            ws_sum.cell(row=sum_row, column=1, value=code).border = thin_border
            ws_sum.cell(row=sum_row, column=2, value=name).border = thin_border

            # 先收集該股所有日期的合計股數
            sum_values = []
            for d in all_dates:
                total = 0
                found = False
                for r in db.execute(
                    "SELECT shares FROM daily_holdings WHERE date=? AND stock_code=?",
                    (d, code),
                ).fetchall():
                    if r[0] is not None:
                        total += r[0]
                        found = True
                sum_values.append(total if found else None)

            for i, val in enumerate(sum_values):
                cell = ws_sum.cell(row=sum_row, column=3 + i, value=val)
                cell.number_format = '#,##0'
                cell.border = thin_border
                if val is None:
                    cell.value = "-"
                    cell.font = Font(color="AAAAAA")
                else:
                    # 找前一個有效值
                    prev_val = None
                    for j in range(i - 1, -1, -1):
                        if sum_values[j] is not None:
                            prev_val = sum_values[j]
                            break
                    if prev_val is not None and val != prev_val:
                        cell.fill = new_fill if val > prev_val else del_fill

            sum_row += 1

        # 欄寬
        ws_sum.column_dimensions['A'].width = 10
        ws_sum.column_dimensions['B'].width = 14
        for i in range(len(all_dates)):
            ws_sum.column_dimensions[get_column_letter(3 + i)].width = 14

    # 存檔
    wb.save(output_path)
    print(f"\n📊 報表已產出：{output_path}")


# ============================================================
#  Main
# ============================================================
def main():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", type=Path, default=None, help="指定單一 xlsx 檔案")
    parser.add_argument("--fund-code", type=str, default=None, help="指定 fund_code（用於舊格式檔名）")
    parser.add_argument("--report", action="store_true", help="只產報表（不重新解析）")
    parser.add_argument("--db", type=Path, default=DB_PATH, help="SQLite 路徑")
    args = parser.parse_args()

    db = sqlite3.connect(args.db)
    init_db(db)

    if not args.report:
        if args.file:
            print(f"--- 解析指定檔案：{args.file} ---")
            parse_xlsx(args.file, db, fund_code_override=args.fund_code)
        else:
            # 舊格式檔名映射（下載腳本已改用 fundCode 前綴，這裡是相容舊檔）
            fc_map = {
                "ETF_Investment_Portfolio": "49YTW",  # 舊格式對應 49YTW
            }
            process_all_xlsx(db, DOWNLOADS_DIR, fund_code_map=fc_map)

    # 產報表
    today_str = datetime.now().strftime("%Y%m%d")
    report_path = REPORT_DIR / f"ETF_Daily_Report_{today_str}.xlsx"
    generate_report(db, report_path)

    db.close()


if __name__ == "__main__":
    main()
