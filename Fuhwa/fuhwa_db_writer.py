"""
復華投信 ETF 資料處理器
- 解析復華下載的 xlsx → 寫入 SQLite（共用 Ezmoney/etf_data.db）
- 跟 Ezmoney 的 etf_processor.py 共用 DB schema

復華 xlsx 格式：
  Row 1:  復華台灣未來50主動式ETF基金...（證劵代碼：00991A）
  Row 3:  日期: 2026/08/12
  Row 5:  基金資產淨值
  Row 6:  90,728,994,922
  Row 8:  基金在外流通單位數
  Row 9:  5,065,916,000
  Row 11: 基金每單位淨值
  Row 12: 17.91
  Row 14: 證券代號 | 證券名稱 | 股數 | 金額 | 權重(%)
  Row 15+: 資料...

使用方式：
  python fuhwa_db_writer.py                        # 處理 downloads/ 下所有 xlsx
  python fuhwa_db_writer.py --file path/to/file.xlsx  # 處理指定檔案
"""

import sqlite3
import re
import sys
import openpyxl
from pathlib import Path
from datetime import datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ====== 路徑設定 ======
FUHWA_DIR = Path(__file__).parent
ETF_ROOT = FUHWA_DIR.parent  # .../ETF
DB_PATH = ETF_ROOT / "Ezmoney" / "etf_data.db"  # 共用 DB
DOWNLOADS_DIR = FUHWA_DIR / "downloads"

# 復華 ETF 代碼 → 上市代碼
ETF_CODE_MAP = {
    "ETF23": "00991A",
}


# ====== 解析工具 ======
def parse_num(val) -> float | None:
    """'90,728,994,922' → 90728994922.0"""
    if val is None:
        return None
    s = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_pct(val) -> float | None:
    """'13.841%' → 13.841"""
    if val is None:
        return None
    s = str(val).replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def parse_date(val) -> str | None:
    """'2026/08/12' → '2026-08-12'"""
    if val is None:
        return None
    s = str(val).strip()
    # Try 2026/08/12
    m = re.match(r"(\d{4})/(\d{2})/(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # Try 2026-08-12
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s
    return None


# ====== DB Schema（跟 Ezmoney 共用） ======
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS daily_summary (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    net_assets  REAL,
    nav         REAL,
    outstanding_units REAL,
    stock_amount REAL,
    cash_amount  REAL,
    futures_margin REAL,
    repo_bonds   REAL,
    receivables  REAL,
    PRIMARY KEY (date, fund_code)
);

CREATE TABLE IF NOT EXISTS daily_holdings (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    stock_code  TEXT NOT NULL,
    stock_name  TEXT NOT NULL,
    shares      INTEGER,
    weight_pct  REAL,
    PRIMARY KEY (date, fund_code, stock_code)
);

CREATE TABLE IF NOT EXISTS daily_futures (
    date        TEXT NOT NULL,
    fund_code   TEXT NOT NULL,
    future_code TEXT NOT NULL,
    future_name TEXT NOT NULL,
    weight_pct  REAL,
    lots        INTEGER,
    contract_month TEXT,
    PRIMARY KEY (date, fund_code, future_code, contract_month)
);
"""


def init_db(db: sqlite3.Connection):
    db.executescript(SCHEMA_SQL)
    db.commit()


# ====== 解析 xlsx → 寫入 DB ======
def parse_xlsx(filepath: Path, db: sqlite3.Connection, fund_code_override: str | None = None) -> bool:
    """解析一個復華 xlsx，寫入 DB。回傳成功與否。"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # --- 從檔名推導 ETF 代碼 ---
    fname = filepath.stem  # e.g. 'ETF23_ETF_Investment_Portfolio_20260812'
    fc_match = re.match(r"^(ETF\d+)_", fname)
    if fc_match:
        etf_id = fc_match.group(1)
        fund_code = ETF_CODE_MAP.get(etf_id, etf_id)
    elif fund_code_override:
        fund_code = fund_code_override
    else:
        fund_code = "00991A"

    # --- 日期：從 Row 3 (A3) 讀取 ---
    raw_date = ws.cell(row=3, column=1).value  # '日期: 2026/08/12'
    if raw_date:
        date_str = parse_date(str(raw_date).replace("日期:", "").replace("日期：", "").strip())
    else:
        date_str = None

    if not date_str:
        print(f"  [ERROR] 無法解析日期: {raw_date}")
        wb.close()
        return False

    # --- 檢查是否已有持股資料 ---
    existing = db.execute(
        "SELECT 1 FROM daily_holdings WHERE date=? AND fund_code=? LIMIT 1",
        (date_str, fund_code),
    ).fetchone()
    if existing:
        print(f"  [SKIP] {fund_code} {date_str} 資料已存在，跳過")
        wb.close()
        return True

    # --- 摘要資料 ---
    net_assets = parse_num(ws.cell(row=6, column=1).value)    # Row 6: 基金資產淨值
    outstanding = parse_num(ws.cell(row=9, column=1).value)     # Row 9: 流通單位數
    nav = parse_num(ws.cell(row=12, column=1).value)           # Row 12: 每單位淨值

    db.execute(
        """INSERT OR REPLACE INTO daily_summary
           (date, fund_code, net_assets, nav, outstanding_units)
           VALUES (?, ?, ?, ?, ?)""",
        (date_str, fund_code, net_assets, nav, outstanding),
    )

    # --- 持股資料：從 Row 15 開始 ---
    holdings_count = 0
    row = 15
    while row <= ws.max_row:
        stock_code = ws.cell(row=row, column=1).value
        stock_name = ws.cell(row=row, column=2).value
        shares = ws.cell(row=row, column=3).value
        amount = ws.cell(row=row, column=4).value
        weight = ws.cell(row=row, column=5).value

        if stock_code is None or str(stock_code).strip() == "":
            break

        stock_code_str = str(stock_code).strip()
        stock_name_str = str(stock_name).strip() if stock_name else ""
        shares_int = int(parse_num(shares) or 0)
        weight_pct = parse_pct(weight)

        db.execute(
            """INSERT OR REPLACE INTO daily_holdings
               (date, fund_code, stock_code, stock_name, shares, weight_pct)
               VALUES (?, ?, ?, ?, ?, ?)""",
            (date_str, fund_code, stock_code_str, stock_name_str, shares_int, weight_pct),
        )
        holdings_count += 1
        row += 1

    db.commit()
    wb.close()

    print(f"  [OK] {fund_code} {date_str} 寫入 {holdings_count} 筆持股")
    return True


def process_dir(db: sqlite3.Connection, dir_path: Path | None = None) -> int:
    """處理目錄下所有 xlsx，回傳成功數。"""
    target = dir_path or DOWNLOADS_DIR
    if not target.exists():
        print(f"  [INFO] 目錄不存在: {target}")
        return 0

    xlsx_files = sorted(target.glob("*.xlsx"))
    if not xlsx_files:
        print(f"  [INFO] 沒有 xlsx 檔案: {target}")
        return 0

    success = 0
    for fp in xlsx_files:
        try:
            if parse_xlsx(fp, db):
                success += 1
        except Exception as e:
            print(f"  [ERROR] {fp.name}: {e}")

    return success


def main():
    import argparse
    parser = argparse.ArgumentParser(description="復華 ETF xlsx → DB")
    parser.add_argument("--file", type=Path, default=None, help="指定 xlsx 檔案")
    parser.add_argument("--dir", type=Path, default=None, help="指定目錄（預設 downloads/）")
    args = parser.parse_args()

    db = sqlite3.connect(str(DB_PATH))
    init_db(db)

    if args.file:
        ok = parse_xlsx(args.file, db)
        print(f"\n結果: {'成功' if ok else '失敗'}")
    else:
        n = process_dir(db, args.dir)
        print(f"\n處理完成: {n} 個檔案成功")

    db.close()


if __name__ == "__main__":
    main()
