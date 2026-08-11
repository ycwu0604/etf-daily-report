"""
跨 ETF 持股合計報表產生器（Capital + Ezmoney 合併版）

資料來源:
  全部從共用 SQLite DB 讀取（Ezmoney/etf_data.db）
  - 49YTW, 63YTW: Ezmoney 下載 xlsx → parse → DB
  - 00982A, 00992A: Capital API → DB（由 Capital/capital_db_writer.py 寫入）

輸出 Excel 5 個 sheet:
  49YTW, 63YTW, 00982A, 00992A, 跨ETF持股合計

每個 ETF sheet 左右兩區：
  左  (A-F)  每日持股比較  (latest vs prev)
  右  (I-...)  歷史股數變化  (按日展開，缺資料顯示 '-')

跨 ETF 持股合計 sheet:
  所有股票代號跨 4 個 ETF 加總，缺資料顯示 '-'，按最新日合計股數降序。

日期採用 UNION（任一 ETF 出現的日期都會出現）。

使用方式:
    python gen_combined_report.py
    python gen_combined_report.py --month 2026-07
    python gen_combined_report.py --from 2026-07-01 --to 2026-07-15

輸出位置:
    {ETF}/reports/ETF_Combined_Daily_Report_{YYYYMMDD}.xlsx

依賴:
    openpyxl, Ezmoney/ezmoney_db_reader.py
"""

import os
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── 動態 import 子模組 ────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))  # .../ETF
_ezmoney_dir = os.path.join(SCRIPT_DIR, 'Ezmoney')

if _ezmoney_dir not in sys.path:
    sys.path.insert(0, _ezmoney_dir)

from ezmoney_db_reader import fetch_all_etf_data, fetch_futures_data, ALL_ETF_CODES  # noqa: E402

# ── 路徑設定 ──────────────────────────────────────────────
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'reports')

# ── ETF 設定 ──────────────────────────────────────────────
SHEET_ORDER = ALL_ETF_CODES  # 49YTW, 63YTW, 00982A, 00992A


# ── 樣式定義 ──────────────────────────────────────────────
BLUE_FILL     = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
GOLD_FILL     = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
ORANGE_FILL   = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
LT_GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LT_RED_FILL   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

WHITE_FONT  = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14)
SUB_FONT    = Font(size=10)
SUB_ITALIC  = Font(size=10, italic=True)
DATA_FONT   = Font(name='Calibri', size=11)
GREEN_FONT  = Font(name='Calibri', size=11, color='006100')
RED_FONT    = Font(name='Calibri', size=11, color='FF0000')
HDR_ALIGN   = Alignment(horizontal='center')


# ── Sheet 建構: 單一 ETF ───────────────────────────────────
def build_etf_sheet(wb, etf_name, etf_data, all_dates, latest_date, prev_date,
                    futures_data=None):
    """建立單一 ETF 的工作表 (左: 每日比較, 右: 歷史股數變化, 期貨持倉)"""
    ws = wb.create_sheet(title=etf_name)

    ws['A1'] = f'{etf_name} 每日持股比較'
    ws['A1'].font = TITLE_FONT
    ws['I1'] = '歷史股數變化'
    ws['I1'].font = TITLE_FONT
    ws['A2'] = f'日期：{latest_date} vs {prev_date}'
    ws['A2'].font = SUB_FONT

    # 左表 header
    for j, h in enumerate(['股票代號', '股票名稱',
                            f'股數 ({prev_date})', f'股數 ({latest_date})',
                            '變化', '變化%'], 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = HDR_ALIGN

    # 右表 header
    for j, h in enumerate(['股票代號', '股票名稱'] + all_dates, 9):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = HDR_ALIGN

    # 收集此 ETF 所有出現過的股票代號
    all_codes = set()
    for d in etf_data:
        all_codes.update(etf_data[d].keys())
    latest_holdings = etf_data.get(latest_date, {})
    sorted_codes = sorted(
        all_codes,
        key=lambda c: (-latest_holdings.get(c, {}).get('shares', 0), c),
    )

    for i, code in enumerate(sorted_codes, 5):
        name = ''
        for d in sorted(etf_data, reverse=True):
            if code in etf_data[d]:
                name = etf_data[d][code].get('name', '') or ''
                if name:
                    break

        ws.cell(row=i, column=1, value=code).font = DATA_FONT
        ws.cell(row=i, column=2, value=name).font = DATA_FONT

        prev_s = etf_data.get(prev_date, {}).get(code, {}).get('shares')
        curr_s = etf_data.get(latest_date, {}).get(code, {}).get('shares')

        if prev_s is not None:
            c = ws.cell(row=i, column=3, value=prev_s)
            c.number_format = '#,##0'
            c.font = DATA_FONT
        else:
            ws.cell(row=i, column=3, value='-').font = DATA_FONT

        if curr_s is not None:
            c = ws.cell(row=i, column=4, value=curr_s)
            c.number_format = '#,##0'
            c.font = DATA_FONT
        else:
            ws.cell(row=i, column=4, value='-').font = DATA_FONT

        if prev_s is not None and curr_s is not None:
            change = curr_s - prev_s
            if change > 0:
                ws.cell(row=i, column=5, value=change).font = GREEN_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                pct = change / prev_s if prev_s != 0 else 0
                ws.cell(row=i, column=6, value=f'{pct:.2%}').font = GREEN_FONT
            elif change < 0:
                ws.cell(row=i, column=5, value=change).font = RED_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                pct = change / prev_s if prev_s != 0 else 0
                ws.cell(row=i, column=6, value=f'{pct:.2%}').font = RED_FONT
            else:
                ws.cell(row=i, column=5, value=0).font = DATA_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                ws.cell(row=i, column=6, value='0.00%').font = DATA_FONT
        elif prev_s is None and curr_s is not None:
            ws.cell(row=i, column=5, value='NEW').font = GREEN_FONT
        elif prev_s is not None and curr_s is None:
            ws.cell(row=i, column=5, value='DEL').font = RED_FONT

        # 右表: 歷史股數
        ws.cell(row=i, column=9, value=code).font = DATA_FONT
        ws.cell(row=i, column=10, value=name).font = DATA_FONT
        for j, d in enumerate(all_dates, 11):
            shares = etf_data.get(d, {}).get(code, {}).get('shares')
            if shares is not None:
                c = ws.cell(row=i, column=j, value=shares)
                c.number_format = '#,##0'
                c.font = DATA_FONT
                date_idx = all_dates.index(d)
                if date_idx > 0:
                    prev_d = all_dates[date_idx - 1]
                    prev_shares = etf_data.get(prev_d, {}).get(code, {}).get('shares')
                    if prev_shares is not None:
                        if shares > prev_shares:
                            c.fill = LT_GREEN_FILL
                        elif shares < prev_shares:
                            c.fill = LT_RED_FILL
            else:
                ws.cell(row=i, column=j, value='-').font = DATA_FONT

    # ── 期貨持倉區塊 ──
    if futures_data and etf_name in futures_data:
        etf_fut = futures_data[etf_name]
        fut_dates = sorted(etf_fut.keys())
        if fut_dates:
            # 找到股票區塊結束 row (最後一個有資料的 row + 2)
            fut_start_row = i + 2 if 'i' in dir() else 5 + len(sorted_codes) + 2
            # 確保不重疊 — 從更下方開始
            for check_row in range(fut_start_row, ws.max_row + 20):
                if ws.cell(row=check_row, column=1).value is None:
                    fut_start_row = check_row
                    break

            ws.cell(row=fut_start_row, column=1, value='期貨持倉').font = TITLE_FONT

            # 期貨表頭: 日期 | 期貨代號 | 期貨名稱 | 持股權重 | 口數 | 契約年月
            hdr_row = fut_start_row + 1
            for j, h in enumerate(['日期', '期貨代號', '期貨名稱', '持股權重', '口數', '契約年月'], 1):
                c = ws.cell(row=hdr_row, column=j, value=h)
                c.fill = ORANGE_FILL
                c.font = WHITE_FONT
                c.alignment = HDR_ALIGN

            data_row = hdr_row + 1
            for d in sorted(etf_fut.keys(), reverse=True):
                for fut in etf_fut[d]:
                    ws.cell(row=data_row, column=1, value=d).font = DATA_FONT
                    ws.cell(row=data_row, column=2, value=fut.get('code', '')).font = DATA_FONT
                    ws.cell(row=data_row, column=3, value=fut.get('name', '')).font = DATA_FONT
                    wt = fut.get('weight_pct')
                    if wt is not None:
                        ws.cell(row=data_row, column=4, value=f'{wt:.2f}%').font = DATA_FONT
                    lots = fut.get('lots')
                    if lots is not None:
                        c = ws.cell(row=data_row, column=5, value=lots)
                        c.number_format = '#,##0'
                        c.font = DATA_FONT
                    ws.cell(row=data_row, column=6, value=fut.get('contract_month', '')).font = DATA_FONT
                    data_row += 1

    # 欄寬
    col_widths = {'A': 10, 'B': 14, 'C': 16, 'D': 16, 'E': 12, 'F': 10,
                   'G': 3, 'H': 3, 'I': 10, 'J': 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for j in range(11, 11 + len(all_dates)):
        ws.column_dimensions[get_column_letter(j)].width = 14


# ── Sheet 建構: 跨 ETF 持股合計 ────────────────────────────
def build_combined_sheet(wb, etf_names, etf_data_list, all_dates, latest_date):
    """建立跨 ETF 持股合計工作表 — 所有股票代號跨 4 個 ETF 加總"""
    ws = wb.create_sheet(title='跨ETF持股合計')

    ws['A1'] = '跨 ETF 持股合計（股數相加）'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'包含：{", ".join(etf_names)}'
    ws['A2'].font = SUB_ITALIC

    for j, h in enumerate(['股票代號', '股票名稱'] + all_dates, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = GOLD_FILL
        c.font = WHITE_FONT
        if j >= 3:
            c.alignment = HDR_ALIGN

    all_codes = set()
    for etf_data in etf_data_list:
        for d in etf_data:
            all_codes.update(etf_data[d].keys())

    def sort_key(code):
        total = sum(
            etf_data.get(latest_date, {}).get(code, {}).get('shares', 0)
            for etf_data in etf_data_list
        )
        return (-total, code)

    sorted_codes = sorted(all_codes, key=sort_key)

    for i, code in enumerate(sorted_codes, 5):
        name = ''
        for etf_data in etf_data_list:
            for d in sorted(etf_data, reverse=True):
                if code in etf_data[d]:
                    cand = etf_data[d][code].get('name', '') or ''
                    if cand:
                        name = cand
                        break
                if name:
                    break
            if name:
                break

        ws.cell(row=i, column=1, value=code).font = DATA_FONT
        ws.cell(row=i, column=2, value=name).font = DATA_FONT

        for j, d in enumerate(all_dates, 3):
            total = 0
            has_data = False
            for etf_data in etf_data_list:
                if code in etf_data.get(d, {}):
                    sh = etf_data[d][code].get('shares')
                    if sh is not None:
                        total += sh
                        has_data = True
            if has_data:
                c = ws.cell(row=i, column=j, value=total)
                c.number_format = '#,##0'
                c.font = DATA_FONT
                date_idx = all_dates.index(d)
                if date_idx > 0:
                    prev_d = all_dates[date_idx - 1]
                    prev_total = 0
                    prev_has = False
                    for etf_data in etf_data_list:
                        if code in etf_data.get(prev_d, {}):
                            psh = etf_data[prev_d][code].get('shares')
                            if psh is not None:
                                prev_total += psh
                                prev_has = True
                    if prev_has:
                        if total > prev_total:
                            c.fill = LT_GREEN_FILL
                        elif total < prev_total:
                            c.fill = LT_RED_FILL
            else:
                ws.cell(row=i, column=j, value='-').font = DATA_FONT

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    for j in range(3, 3 + len(all_dates)):
        ws.column_dimensions[get_column_letter(j)].width = 14


# ── 命令列參數解析 ─────────────────────────────────────────
def parse_args():
    args = sys.argv[1:]
    month = None
    date_from = None
    date_to = None

    i = 0
    while i < len(args):
        if args[i] == '--month' and i + 1 < len(args):
            month = args[i + 1]
            i += 2
        elif args[i] == '--from' and i + 1 < len(args):
            date_from = args[i + 1]
            i += 2
        elif args[i] == '--to' and i + 1 < len(args):
            date_to = args[i + 1]
            i += 2
        else:
            i += 1

    if month and not date_from and not date_to:
        date_from = f'{month}-01'
        y, m = int(month[:4]), int(month[5:7])
        if m == 12:
            last_day = 31
        else:
            next_month = datetime(y, m + 1, 1) - timedelta(days=1)
            last_day = next_month.day
        date_to = f'{month}-{last_day:02d}'

    return date_from, date_to


# ── 主程式 ────────────────────────────────────────────────
def main():
    date_from, date_to = parse_args()

    if date_from and date_to:
        print(f'Date range: {date_from} ~ {date_to}')
    else:
        today = datetime.now().strftime('%Y-%m-%d')
        first_of_month = datetime.now().replace(day=1).strftime('%Y-%m-%d')
        print(f'Date range: (auto) {first_of_month} ~ {today}')

    # 全部從 DB 讀取
    print('\n[DB] Reading all ETF data from SQLite...')
    all_etf_data = fetch_all_etf_data(date_from, date_to)

    # 讀取期貨資料
    print('\n[DB] Reading futures data from SQLite...')
    futures_data = fetch_futures_data(date_from=date_from, date_to=date_to)

    # 過濾: 沒有任何資料的 ETF
    for fc in list(all_etf_data.keys()):
        if not all_etf_data[fc]:
            print(f'  [WARN] {fc}: 沒有任何資料，略過此 sheet')

    # 統整所有日期 (UNION)
    all_dates = sorted({d for data in all_etf_data.values() for d in data})
    if not all_dates:
        print('[ERROR] No trading days found in any source, exiting.')
        return

    latest_date = all_dates[-1]
    prev_date = all_dates[-2] if len(all_dates) >= 2 else all_dates[-1]
    print(f'\nDate range: {all_dates[0]} ~ {all_dates[-1]}')
    print(f'Latest: {latest_date}, Previous: {prev_date}')

    # 建立 Workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    etf_names = []
    etf_data_list = []
    for etf_code in SHEET_ORDER:
        if etf_code not in all_etf_data or not all_etf_data[etf_code]:
            continue
        build_etf_sheet(
            wb, etf_code, all_etf_data[etf_code],
            all_dates, latest_date, prev_date,
            futures_data=futures_data,
        )
        etf_names.append(etf_code)
        etf_data_list.append(all_etf_data[etf_code])

    if len(etf_data_list) >= 2:
        build_combined_sheet(wb, etf_names, etf_data_list, all_dates, latest_date)

    # 儲存
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    latest_yyyymmdd = latest_date.replace('-', '')
    output_path = os.path.join(OUTPUT_DIR, f'ETF_Combined_Daily_Report_{latest_yyyymmdd}.xlsx')
    try:
        with open(output_path, 'a'):
            pass
    except (PermissionError, OSError):
        ts = datetime.now().strftime('%H%M%S')
        output_path = os.path.join(OUTPUT_DIR, f'ETF_Combined_Daily_Report_{latest_yyyymmdd}_{ts}.xlsx')
        print(f'[INFO] Original file locked, saving as: {os.path.basename(output_path)}')
    wb.save(output_path)
    print(f'\n[OK] Report saved: {output_path}')
    print(f'   Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
