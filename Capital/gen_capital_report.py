"""
Capital ETF 每日持股報告產生器 (API 版)

從群益投信 CFWeb API 自動下載每日持股資料，
產生包含「每日持股比較 + 歷史股數變化 + 跨 ETF 持股合計」的報告 Excel。

API 端點: POST https://www.capitalfund.com.tw/CFWeb/api/etf/buyback
  body: {"fundId": "<id>", "date": "YYYY/MM/DD" | null}

日期對應: API 回傳的 date1 是「申購買回日」，持股資料實際是 date1 前一交易日 (=date2)。
  因此報告中的日期使用 date2（代表實際持股日）。

使用方式:
    python gen_capital_report.py              # API 模式（自動下載當月所有交易日資料）
    python gen_capital_report.py --mode excel # Excel 模式（讀取本地 00982A/00992A 資料夾）
    python gen_capital_report.py --month 2026-06  # 指定年月
    python gen_capital_report.py --from 2026-07-01 --to 2026-07-15  # 指定日期範圍

輸出位置:
    {BASE_DIR}/reports/ETF_Daily_Report_{最新日期}.xlsx

依賴:
    pip install openpyxl requests
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import sys
import json
import ssl
import urllib.request
from datetime import datetime, timedelta
from http.cookiejar import CookieJar

# ── 路徑設定 ──────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = SCRIPT_DIR
OUTPUT_DIR = os.path.join(SCRIPT_DIR, 'reports')

# ── ETF 設定 ──────────────────────────────────────────────
# fundId 對應: 網址用 399/400，但 API 用 399/500
ETF_CONFIG = {
    '00982A': {'fundId': '399', 'name': '主動群益台灣強棒'},
    '00992A': {'fundId': '500', 'name': '主動群益科技創新'},
}
ETF_DIRS = list(ETF_CONFIG.keys())

# ── API 設定 ──────────────────────────────────────────────
API_BASE = 'https://www.capitalfund.com.tw/CFWeb/api/etf/buyback'
API_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'application/json, text/plain, */*',
    'Content-Type': 'application/json',
    'Referer': 'https://www.capitalfund.com.tw/etf/product/detail/399/portfolio',
    'Origin': 'https://www.capitalfund.com.tw',
}

# ── 年月設定 (Excel 模式用) ──────────────────────────────
YEAR_MONTH = None  # None = 自動偵測當月


# ── SSL Context ────────────────────────────────────────────
def _ssl_ctx():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx


# ── API 呼叫 ──────────────────────────────────────────────
def call_buyback_api(fund_id, date_val=None, retries=3):
    """
    呼叫 CFWeb buyback API，含自動重試。
    date_val: None (最新) 或 "YYYY/MM/DD" 或 "YYYY-MM-DD"
    回傳: API JSON (含 pcf, stocks, bonds 等)
    """
    body = json.dumps({"fundId": str(fund_id), "date": date_val}).encode('utf-8')
    last_err = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(API_BASE, data=body, headers=API_HEADERS, method='POST')
            resp = urllib.request.urlopen(req, timeout=20, context=_ssl_ctx())
            return json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            last_err = e
    raise last_err


# ── 工具函式 ──────────────────────────────────────────────
def parse_shares(val):
    """將含逗號的股數字串轉為 int，無法轉換則回傳 None"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('-', '(無)', ''):
        return None
    s = s.replace(',', '').replace(',', '')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def detect_year_month_excel(etf_dir):
    """從 ETF 資料夾的 Excel 檔名推斷年月 (民國年/西元年+月) — Excel 模式用"""
    base = os.path.join(BASE_DIR, etf_dir)
    files = [f for f in os.listdir(base) if f.endswith('.xlsx')]
    if not files:
        return None
    path = os.path.join(base, files[0])
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=2, values_only=True):
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell)
            m = re.search(r'(\d{3})/(\d{2})/(\d{2})', cell_str)
            if m:
                roc_year = int(m.group(1))
                month = m.group(2)
                west_year = roc_year + 1911
                return f'{west_year}-{month}'
    wb.close()
    return None


# ── API 模式: 下載持股資料 ────────────────────────────────
def fetch_etf_data_api(etf_code, date_from=None, date_to=None):
    """
    透過 API 下載 ETF 在指定期間內所有交易日的持股資料。
    
    回傳: {date_str: {code: {'name': str, 'shares': int}}}
    其中 date_str 是 date2 (前一交易日 = 實際持股日)，格式 YYYY-MM-DD
    
    策略: 逐日查詢 date_from~date_to 範圍，跳過無資料的日期（非交易日）。
    若 date_from/date_to 未指定，預設為當月 1 號~今天。
    """
    fund_id = ETF_CONFIG[etf_code]['fundId']
    today = datetime.now().strftime('%Y-%m-%d')
    
    # 預設: 當月
    if date_from is None:
        date_from = datetime.now().replace(day=1).strftime('%Y-%m-%d')
    if date_to is None:
        date_to = today
    
    print(f'  [API] {etf_code} (fundId={fund_id}): fetching {date_from} ~ {date_to}')
    
    all_data = {}
    seen_dates = set()  # 避免重複 (date2 可能被多個 date1 指到)
    
    # 逐日掃描
    start = datetime.strptime(date_from, '%Y-%m-%d')
    end = datetime.strptime(date_to, '%Y-%m-%d')
    
    # 為了涵蓋 date_to 的持股，需要查詢 date_to+1 (因 date2=date1-1)
    # 遇週末/連假時 API 回傳空，需多查數天才能取到 date_to 的持股
    # 例: 7/31(五) 持股 → date1=8/3(一)，需多查 3 天
    query_end = end + timedelta(days=5)
    
    current = start
    while current <= query_end:
        date_str = current.strftime('%Y/%m/%d')
        try:
            result = call_buyback_api(fund_id, date_str)
            data = result.get('data', {})
            stocks = data.get('stocks', []) if isinstance(data, dict) else []
            pcf = data.get('pcf', {}) if isinstance(data, dict) else {}
            date1 = pcf.get('date1', '')
            date2 = pcf.get('date2', '')
            
            if stocks and date2:
                # date2 是實際持股日; 只收在指定範圍內的日期
                if date2 not in seen_dates and date_from <= date2 <= date_to:
                    seen_dates.add(date2)
                    holdings = {}
                    for s in stocks:
                        code = str(s.get('stocNo', '')).strip()
                        name = s.get('stocName', '')
                        shares = parse_shares(s.get('shareFormat', ''))
                        if code and shares is not None:
                            holdings[code] = {'name': name, 'shares': shares}
                    all_data[date2] = holdings
        except Exception as e:
            # 非交易日 (HTTP 500) 或網路逾時 — 跳過
            print(f'    [WARN] query={date_str} failed: {type(e).__name__}')
            pass
        current += timedelta(days=1)
    
    return all_data


# ── Excel 模式: 讀取本地檔案 ──────────────────────────────
def read_etf_data_excel(etf_dir, year_month=None):
    """讀取 ETF 資料夾中所有 xlsx 的第二個 sheet，回傳 {date_str: {code: {name, shares}}}"""
    base = os.path.join(BASE_DIR, etf_dir)
    files = sorted([f for f in os.listdir(base) if f.endswith('.xlsx')])
    all_data = {}

    ym = year_month or YEAR_MONTH
    if ym is None:
        ym = detect_year_month_excel(etf_dir)

    for f in files:
        m = re.search(r'_(\d{4})\.xlsx$', f)
        if not m:
            continue
        ds = m.group(1)
        if ym:
            date_str = f'{ym}-{ds[2:]}'
        else:
            date_str = f'{ds[:2]}-{ds[2:]}'

        path = os.path.join(base, f)
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[1]]
        holdings = {}
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            code = str(row[0].value).strip() if row[0].value else None
            name = row[1].value
            shares = parse_shares(row[3].value)
            if code and shares is not None:
                holdings[code] = {'name': name, 'shares': shares}
        all_data[date_str] = holdings
        wb.close()
    return all_data


# ── 樣式定義 ──────────────────────────────────────────────
BLUE_FILL     = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
GOLD_FILL     = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
LT_GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LT_RED_FILL   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

WHITE_FONT  = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT  = Font(bold=True, size=14)
SUB_FONT    = Font(size=10)
DATA_FONT   = Font(name='Calibri', size=11)
GREEN_FONT  = Font(name='Calibri', size=11, color='006100')
RED_FONT    = Font(name='Calibri', size=11, color='FF0000')
HDR_ALIGN   = Alignment(horizontal='center')


# ── Sheet 建構 ─────────────────────────────────────────────
def build_etf_sheet(wb, etf_name, etf_data, all_dates, latest_date, prev_date):
    """建立單一 ETF 的工作表 (左: 每日比較, 右: 歷史股數變化)"""
    ws = wb.create_sheet(title=etf_name)

    ws['A1'] = f'{etf_name} 每日持股比較'
    ws['A1'].font = TITLE_FONT
    ws['I1'] = '歷史股數變化'
    ws['I1'].font = TITLE_FONT
    ws['A2'] = f'日期: {latest_date} vs {prev_date}'
    ws['A2'].font = SUB_FONT

    # 左表 header
    for j, h in enumerate(['股票代號', '股票名稱',
                            f'股數 ({prev_date})', f'股數 ({latest_date})',
                            '變化', '變化%'], 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = HDR_ALIGN

    # 右表 header
    for j, h in enumerate(['股票代號', '股票名稱'] + all_dates, 9):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = BLUE_FILL
        c.font = WHITE_FONT
        c.alignment = HDR_ALIGN

    latest_holdings = etf_data.get(latest_date, {})
    all_codes = set()
    for d in etf_data:
        all_codes.update(etf_data[d].keys())
    sorted_codes = sorted(all_codes, key=lambda c: (-latest_holdings.get(c, {}).get('shares', 0), c))

    for i, code in enumerate(sorted_codes, 5):
        name = ''
        for d in sorted(etf_data, reverse=True):
            if code in etf_data[d]:
                name = etf_data[d][code]['name']
                break

        ws.cell(row=i, column=1, value=code).font = DATA_FONT
        ws.cell(row=i, column=2, value=name).font = DATA_FONT

        prev_s = etf_data.get(prev_date, {}).get(code, {}).get('shares')
        curr_s = etf_data.get(latest_date, {}).get(code, {}).get('shares')

        if prev_s is not None:
            c = ws.cell(row=i, column=3, value=prev_s)
            c.number_format = '#,##0'
            c.font = DATA_FONT
        else:
            ws.cell(row=i, column=3, value='(無)').font = DATA_FONT

        if curr_s is not None:
            c = ws.cell(row=i, column=4, value=curr_s)
            c.number_format = '#,##0'
            c.font = DATA_FONT
        else:
            ws.cell(row=i, column=4, value='(無)').font = DATA_FONT

        if prev_s is not None and curr_s is not None:
            change = curr_s - prev_s
            if change > 0:
                ws.cell(row=i, column=5, value=change).font = GREEN_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                pct = change / prev_s if prev_s != 0 else 0
                ws.cell(row=i, column=6, value=f'{pct:.2%}').font = GREEN_FONT
            elif change < 0:
                ws.cell(row=i, column=5, value=change).font = RED_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                pct = change / prev_s if prev_s != 0 else 0
                ws.cell(row=i, column=6, value=f'{pct:.2%}').font = RED_FONT
            else:
                ws.cell(row=i, column=5, value=0).font = DATA_FONT
                ws.cell(row=i, column=5).number_format = '#,##0'
                ws.cell(row=i, column=6, value='0.00%').font = DATA_FONT
        elif prev_s is None and curr_s is not None:
            ws.cell(row=i, column=5, value='NEW').font = GREEN_FONT
        elif prev_s is not None and curr_s is None:
            ws.cell(row=i, column=5, value='DEL').font = RED_FONT

        # 右表: 歷史股數
        ws.cell(row=i, column=9, value=code).font = DATA_FONT
        ws.cell(row=i, column=10, value=name).font = DATA_FONT
        for j, d in enumerate(all_dates, 11):
            shares = etf_data.get(d, {}).get(code, {}).get('shares')
            if shares is not None:
                c = ws.cell(row=i, column=j, value=shares)
                c.number_format = '#,##0'
                c.font = DATA_FONT
                date_idx = all_dates.index(d)
                if date_idx > 0:
                    prev_d = all_dates[date_idx - 1]
                    prev_shares = etf_data.get(prev_d, {}).get(code, {}).get('shares')
                    if prev_shares is not None:
                        if shares > prev_shares:
                            c.fill = LT_GREEN_FILL
                        elif shares < prev_shares:
                            c.fill = LT_RED_FILL
            else:
                ws.cell(row=i, column=j, value='-').font = DATA_FONT

    col_widths = {'A': 10, 'B': 14, 'C': 16, 'D': 16, 'E': 12, 'F': 10,
                   'G': 3, 'H': 3, 'I': 10, 'J': 14}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w
    for j in range(11, 11 + len(all_dates)):
        ws.column_dimensions[get_column_letter(j)].width = 14

    return ws


def build_combined_sheet(wb, etf_names, etf_data_list, all_dates, latest_date):
    """建立跨 ETF 持股合計工作表"""
    ws = wb.create_sheet(title='跨ETF持股合計')
    ws['A1'] = '跨 ETF 持股合計（股數相加）'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'包含: {", ".join(etf_names)}'
    ws['A2'].font = SUB_FONT

    for j, h in enumerate(['股票代號', '股票名稱'] + all_dates, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.fill = GOLD_FILL
        c.font = WHITE_FONT
        if j >= 3:
            c.alignment = HDR_ALIGN

    all_codes = set()
    for etf_data in etf_data_list:
        for d in etf_data:
            all_codes.update(etf_data[d].keys())

    def sort_key(code):
        total = sum(etf_data.get(latest_date, {}).get(code, {}).get('shares', 0)
                    for etf_data in etf_data_list)
        return (-total, code)

    sorted_codes = sorted(all_codes, key=sort_key)

    for i, code in enumerate(sorted_codes, 5):
        name = ''
        for etf_data in etf_data_list:
            for d in sorted(etf_data, reverse=True):
                if code in etf_data[d]:
                    name = etf_data[d][code]['name']
                    break
            if name:
                break

        ws.cell(row=i, column=1, value=code).font = DATA_FONT
        ws.cell(row=i, column=2, value=name).font = DATA_FONT

        for j, d in enumerate(all_dates, 3):
            total = 0
            has_data = False
            for etf_data in etf_data_list:
                if code in etf_data.get(d, {}):
                    total += etf_data[d][code]['shares']
                    has_data = True
            if has_data:
                c = ws.cell(row=i, column=j, value=total)
                c.number_format = '#,##0'
                c.font = DATA_FONT
                date_idx = all_dates.index(d)
                if date_idx > 0:
                    prev_d = all_dates[date_idx - 1]
                    prev_total = 0
                    prev_has = False
                    for etf_data in etf_data_list:
                        if code in etf_data.get(prev_d, {}):
                            prev_total += etf_data[prev_d][code]['shares']
                            prev_has = True
                    if prev_has:
                        if total > prev_total:
                            c.fill = LT_GREEN_FILL
                        elif total < prev_total:
                            c.fill = LT_RED_FILL
            else:
                ws.cell(row=i, column=j, value='-').font = DATA_FONT

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    for j in range(3, 3 + len(all_dates)):
        ws.column_dimensions[get_column_letter(j)].width = 14

    return ws


# ── 命令列參數解析 ─────────────────────────────────────────
def parse_args():
    args = sys.argv[1:]
    mode = 'api'  # 預設 API 模式
    month = None
    date_from = None
    date_to = None

    i = 0
    while i < len(args):
        if args[i] == '--mode' and i + 1 < len(args):
            mode = args[i + 1]
            i += 2
        elif args[i] == '--month' and i + 1 < len(args):
            month = args[i + 1]
            i += 2
        elif args[i] == '--from' and i + 1 < len(args):
            date_from = args[i + 1]
            i += 2
        elif args[i] == '--to' and i + 1 < len(args):
            date_to = args[i + 1]
            i += 2
        else:
            i += 1

    # --month 轉成 date_from/date_to
    if month and not date_from and not date_to:
        date_from = f'{month}-01'
        # date_to: 該月最後一天
        y, m = int(month[:4]), int(month[5:7])
        if m == 12:
            last_day = 31
        else:
            next_month = datetime(y, m + 1, 1) - timedelta(days=1)
            last_day = next_month.day
        date_to = f'{month}-{last_day:02d}'

    return mode, date_from, date_to


# ── 主程式 ────────────────────────────────────────────────
def main():
    mode, date_from, date_to = parse_args()

    print(f'Mode: {mode.upper()}')
    if mode == 'api':
        print(f'Date range: {date_from or "(auto: current month)"} ~ {date_to or "(auto: today)"}')

    all_etf_data = {}

    if mode == 'api':
        # API 模式
        for etf_code in ETF_DIRS:
            data = fetch_etf_data_api(etf_code, date_from, date_to)
            all_etf_data[etf_code] = data
            print(f'  {etf_code}: {len(data)} dates, {sorted(data.keys())}')

    elif mode == 'excel':
        # Excel 模式 (向後相容)
        for etf_dir in ETF_DIRS:
            etf_path = os.path.join(BASE_DIR, etf_dir)
            if not os.path.isdir(etf_path):
                print(f'  [SKIP] Dir not found: {etf_dir}')
                continue
            data = read_etf_data_excel(etf_dir)
            all_etf_data[etf_dir] = data
            print(f'  {etf_dir}: {len(data)} dates, {sorted(data.keys())}')

    else:
        print(f'[ERROR] Unknown mode: {mode}. Use "api" or "excel".')
        return

    if not all_etf_data:
        print('[ERROR] No ETF data found, exiting.')
        return

    # 統整所有日期
    all_dates = sorted(set(d for data in all_etf_data.values() for d in data))
    if not all_dates:
        print('[ERROR] No trading days found, exiting.')
        return
    latest_date = all_dates[-1]
    prev_date = all_dates[-2] if len(all_dates) >= 2 else all_dates[-1]
    print(f'Date range: {all_dates[0]} ~ {all_dates[-1]}')
    print(f'Latest: {latest_date}, Previous: {prev_date}')

    # 建立 Workbook
    wb = openpyxl.Workbook()

    etf_names = []
    etf_data_list = []
    for etf_code in ETF_DIRS:
        if etf_code not in all_etf_data:
            continue
        etf_data = all_etf_data[etf_code]
        build_etf_sheet(wb, etf_code, etf_data, all_dates, latest_date, prev_date)
        etf_names.append(etf_code)
        etf_data_list.append(etf_data)

    if len(etf_data_list) >= 2:
        build_combined_sheet(wb, etf_names, etf_data_list, all_dates, latest_date)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # 儲存
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, f'ETF_Daily_Report_{latest_date.replace("-", "")}.xlsx')
    try:
        with open(output_path, 'a'):
            pass
    except (PermissionError, OSError):
        ts = datetime.now().strftime('%H%M%S')
        output_path = os.path.join(OUTPUT_DIR, f'ETF_Daily_Report_{latest_date.replace("-", "")}_{ts}.xlsx')
        print(f'[INFO] Original file locked, saving as: {os.path.basename(output_path)}')
    wb.save(output_path)
    print(f'\n[OK] Report saved: {output_path}')
    print(f'   Sheets: {wb.sheetnames}')


if __name__ == '__main__':
    main()
